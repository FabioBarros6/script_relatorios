import pandas as pd
import openpyxl 
import re
import os

caminho_txt = '' # Colocar entre as aspas o caminho do arquivo TXT onde estão os dados
caminho_planilha = '' # Colocar entre as aspas o caminho da planilha onde serão inseridos os dados contidos no arquivo TXT

def processar_arquivo_txt(caminho_txt, caminho_planilha):
    # Ler o arquivo
    with open(caminho_txt, 'r', encoding='utf-8') as f:
        linhas = f.readlines()
    
    dados = []
    for linha in linhas:
        # Remover os parênteses
        linha = linha.strip().strip('()`')
        
        # Separar os pares chave: valor
        pares = re.split(r',\s*', linha)
        dados_dict = {}
        
        for par in pares:
            if ':' in par:
                chave, valor = map(str.strip, par.split(':', 1))
                dados_dict[chave] = valor
        
        dados.append(dados_dict)
    
    # Criar um DataFrame do pandas
    df = pd.DataFrame(dados)
    
    # Remover colunas duplicadas
    df = df.loc[:, ~df.columns.duplicated()]
    
    if os.path.exists(caminho_planilha):
        # Abre uma planilha existente e adiciona os dados
        book = openpyxl.load_workbook(caminho_planilha)
        if 'Dados' not in book.sheetnames:
            sheet = book.create_sheet('Dados')
            sheet.append(df.columns.tolist())  # Adicionar cabeçalhos
        else:
            sheet = book['Dados']
        
        for row in df.itertuples(index=False, name=None):
            sheet.append(row)
        
        book.save(caminho_planilha)
        book.close()
    else:
        # Cria uma nova planilha e adiciona os dados
        df.to_excel(caminho_planilha, sheet_name='Dados', index=False)
    
    print(f'Dados inseridos na planilha: {caminho_planilha}')

processar_arquivo_txt(caminho_txt, caminho_planilha)